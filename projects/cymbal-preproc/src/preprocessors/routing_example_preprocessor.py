import os
from typing import Any, Dict, List, Optional, Type
from openpyxl import load_workbook
from pydantic import BaseModel, Extra
from google.cloud import storage
from langchain.embeddings.base import Embeddings
from langchain.prompts.example_selector.base import BaseExampleSelector
from langchain.vectorstores.base import VectorStore
import logging
from pathlib import Path
import datetime

LOGGER = logging.getLogger(__name__)
#FIELDS = ["Query", "Intent", "Response"]

def get_sorted_values(example: Dict[str, str]) -> List[str]:
    """Function to sort the values in a dictionary. This is an assumed function, modify as necessary."""
    return sorted(example.values())

class SemanticSimilarityExamplePreprocessor(BaseExampleSelector, BaseModel):
    vectorstore: VectorStore
    k: int = 3
    fetch_k: int = 20
    example_keys: Optional[List[str]] = None
    input_keys: Optional[List[str]] = None

    class Config:
        extra = Extra.forbid
        arbitrary_types_allowed = True
    
    class Config:
        """Configuration for this pydantic object."""

        extra = Extra.forbid
        arbitrary_types_allowed = True

    def add_example(self, example: Dict[str, str]) -> str:
        """Add new example to vectorstore."""
        if self.input_keys:
            string_example = " ".join(
                sorted_values({key: example[key] for key in self.input_keys})
            )
        else:
            string_example = " ".join(sorted_values(example))
        ids = self.vectorstore.add_texts([string_example], metadatas=[example])
        return ids[0]


    def select_examples(self, input_variables: Dict[str, str]) -> List[dict]:
        """Select which examples to use based on semantic similarity."""
        if self.input_keys:
            input_variables = {key: input_variables[key] for key in self.input_keys}
        query = " ".join(sorted_values(input_variables))
        example_docs = self.vectorstore.similarity_search(query, k=self.k)
        examples = [dict(e.metadata) for e in example_docs]
        if self.example_keys:
            examples = [{k: eg[k] for k in self.example_keys} for eg in examples]
        return examples

    @staticmethod
    def load_examples_from_path(path: str, routing_example_fileheader: List[str] = ["Query", "Intent", "Response"]) -> List[Dict[str, str]]:
        workbook = load_workbook(filename=path, data_only=True)  # Set data_only=True to load values only
        return [
            dict(zip(routing_example_fileheader, row)) 
            for sheet in workbook.sheetnames 
            for row in workbook[sheet].iter_rows(min_row=2, values_only=True)
        ]

    @staticmethod
    def download_blob(bucket_name: str, source_blob_name: str, destination_file_path: str) -> None:
        try:
            storage_client = storage.Client()
            bucket = storage_client.bucket(bucket_name)
            blob = bucket.blob(source_blob_name)
            blob.download_to_filename(destination_file_path)
        except Exception as e:
            LOGGER.error(f"Failed to download blob: {e}")
            raise
    
    @staticmethod
    def copy_from_local_to_gcs(bucket_name: str, local_folder_name: str):
        storage_client = storage.Client()
        bucket = storage_client.bucket(bucket_name)
        local_folder_path = os.path.abspath(local_folder_name)
        for root, dirs, files in os.walk(local_folder_path):
            for file in files:
                gcs_file_name = os.path.join(local_folder_name, str(os.path.basename(file)))
                local_file_name = os.path.join(root, file)
                blob = bucket.blob(gcs_file_name)
                blob.upload_from_filename(local_file_name)
                logging.info(f"Uploaded {local_file_name} to {gcs_file_name}")
            
    @classmethod
    def from_examples(
        cls,
        examples: List[Dict[str, str]],
        embeddings: Embeddings,
        vectorstore_cls: Type[VectorStore],
        k: int = 4,
        fetch_k: int = 20,
        bucket_name: str = "gs://ioh-kb-bucket",
        kdb_path: str = "raw_input",
        input_keys: Optional[List[str]] = None,
        index_name: Optional[str] = "similarity_example_selector_index",
        index_output_dir: str = "preprocessed_output",
        **vectorstore_cls_kwargs: Any,
    ) -> "SemanticSimilarityExamplePreprocessor":
        if input_keys:
            string_examples = [
                " ".join(get_sorted_values({k: eg.get(k, '') for k in input_keys}))
                for eg in examples
            ]
        else:
            string_examples = [" ".join(get_sorted_values(eg)) for eg in examples]
        
        index_output_path = Path(index_output_dir) / index_name
        try:
            vectorstore = vectorstore_cls.from_texts(
                string_examples,
                embeddings,
                metadatas=examples,
                **vectorstore_cls_kwargs,
            )
            vectorstore.save_local(index_output_path.as_posix())
            logging.info("Saved Index Locally")
            cls.copy_from_local_to_gcs(bucket_name, index_output_path.as_posix())
            return cls(vectorstore=vectorstore, k=k, fetch_k=fetch_k, input_keys=input_keys)
        except Exception as e:
            logging.exception(f"Failed to create and save the index: {e}")

    @classmethod
    def load_and_process_data(
        cls,
        embeddings: Embeddings,
        vectorstore_cls: Type[VectorStore],
        k: int = 5,
        fetch_k: int = 20,
        bucket_name: str = "gs://ioh-kb-bucket",
        kdb_path: str = "raw_input",
        routing_example_filename: str = "nlu-router-examples.xlsx",
        routing_example_fileheader: List[str] = ["Query", "Intent", "Response"],
        input_keys: Optional[List[str]] = None,
        index_name: str = "nlp_router_examples",
        index_output_dir: str = "preprocessed_output"
    ) -> None:
        print(f"Loading data from {routing_example_filename} with Header Information {routing_example_fileheader} for Index {index_name}")
        folder_path = Path(kdb_path)
        if not folder_path.exists():
            folder_path.mkdir(parents=True)
            logging.info(f"Folder created: {folder_path}")
        else:
            logging.info(f"Folder already exists: {folder_path}")
        
        routing_example_file = folder_path / routing_example_filename
        bucket_name = bucket_name[5:] if bucket_name.startswith("gs://") else bucket_name
        blob_path = Path(kdb_path) / routing_example_filename
        try:
            cls.download_blob(bucket_name, blob_path.as_posix(), routing_example_file)
        except Exception as e:
            logging.exception(f"Failed to download blob: {e}")

        try:
            all_examples = cls.load_examples_from_path(routing_example_file, routing_example_fileheader)
            print(f"Generating embeddings for {len(all_examples)} routing examples. This will take some time. STARTED {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            cls.from_examples(
                all_examples,
                embeddings,
                vectorstore_cls=vectorstore_cls,
                k=k,
                fetch_k=fetch_k,
                bucket_name=bucket_name[5:] if "gs://" in bucket_name else bucket_name,
                kdb_path = kdb_path,
                input_keys=input_keys,
                index_name=index_name,
                index_output_dir= index_output_dir

            )
            print(f"Generated embeddings for {len(all_examples)} routing examples. COMPLETED {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"Generated Index {index_name}")
        except Exception as e:
            LOGGER.exception("Failed to load and process data from sheet")
            raise
